import math
from tqdm import tqdm
import matplotlib.pyplot as plt
import zipfile
from openpyxl import Workbook

def distance_cal(cor1, cor2):
    x1, z1 = cor1
    x2, z2 = cor2
    return math.sqrt((x2 - x1) ** 2 + (z2 - z1) ** 2)

def generate_path(start_point, corlist, num_path):
    path = [start_point]

    with tqdm(total=num_path - 1) as pbar:
        for _ in range(1, num_path):
            min_distance = float('inf')
            min_index = -1
            for i, cor in enumerate(corlist):
                dist = distance_cal(path[-1], cor)
                if dist < min_distance:
                    min_distance = dist
                    min_index = i

            path.append(corlist.pop(min_index))
            pbar.update(1)

    return path

def plot_points_and_path(corlist, path, start_point):
    x_values, y_values = zip(*corlist)
    path_x_values, path_y_values = zip(*path)

    plt.plot(x_values, y_values, 'o', color='lightgray', label='All Points')
    plt.plot(path_x_values, path_y_values, '-o', color='red', label='Path')
    plt.plot(start_point[0], start_point[1], 'o', color='purple', label='Start Point')
    plt.xlabel('X')
    plt.ylabel('Y')
    plt.title('Path')
    plt.grid(True)
    plt.legend()
    plt.show()

def write_waypoints_to_file(waypoints, filename):
    with open(filename, "w") as output_file:
        for counter, x in enumerate(waypoints, 1):
            output_file.write(f"waypoint:{counter}:{counter}:{x[0]}:100:{x[1]}:11:false:0:gui.xaero_default:false:0:false\n")

def create_index_excel(files):
    wb = Workbook()
    ws = wb.active
    ws.title = "Index"

    ws.cell(row=1, column=1, value="File Name")
    ws.cell(row=1, column=2, value="Number of Waypoints")

    for i, file in enumerate(files, 2):
        ws.cell(row=i, column=1, value=file)
        ws.cell(row=i, column=2, value=len(files[file]))

    return wb

def main():
    startx = input("Please input the X coordinate of the start point: ")
    startz = input("Please input the Z coordinate of the start point: ")

    start_point = [int(startx) if startx else 0, int(startz) if startz else 0]

    try:
        num_path = int(input("Please input the number of paths you want to explore: "))
    except ValueError:
        print("Invalid input. Please enter a valid number.")
        exit(1)

    shipOrNotIn = input("Please input 'Y' if you want to explore the ship, or 'N' otherwise: ")
    shipOrNot = shipOrNotIn.upper() == "Y"

    corlist = []

    with open("file.txt") as file:
        file.readline()
        for line in file:
            if "end_city" in line:
                seed, type, x, y, details = line.split(";", 4)
                x, y = int(x), int(y)
                if shipOrNot and "ship" in details or not shipOrNot:
                    corlist.append([x, y])
                    if shipOrNot:
                        print(line)

    path = generate_path(start_point, corlist, num_path)

    total_distance = sum(distance_cal(path[i], path[i+1]) for i in range(len(path)-1))
    distance_ratio = total_distance / num_path
    print(f"Total Distance: {total_distance:.2f}")
    print(f"Distance Ratio: {distance_ratio:.2f}")

    num_files = len(path) // 100
    if len(path) % 100 != 0:
        num_files += 1

    files = {}
    for i in range(num_files):
        waypoints = path[i * 100: (i + 1) * 100]
        filename = f"output_{i+1}.txt"
        files[filename] = waypoints
        write_waypoints_to_file(waypoints, filename)

    with zipfile.ZipFile("output.zip", "w") as zipf:
        for file in files:
            zipf.write(file)

    index_wb = create_index_excel(files)
    index_wb.save("index.xlsx")

    plot_points_and_path(corlist, path, start_point)

if __name__ == "__main__":
    main()
