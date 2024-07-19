import math
import time
import zipfile
import os
from tqdm import tqdm
import matplotlib.pyplot as plt
from openpyxl import Workbook
import generate_path_module


def calculate_distance_no_sqrt(cor1, cor2):
    x1, z1 = cor1
    x2, z2 = cor2
    return ((x2 - x1) ** 2 + (z2 - z1) ** 2)


def generate_path(start_point, coordinates_list, num_paths):
    path = [start_point]
    for _ in range(1, num_paths):
        min_distance = float('inf')
        min_index = -1
        for i, cor in enumerate(coordinates_list):
            dist = calculate_distance_no_sqrt(path[-1], cor)
            if dist < min_distance:
                min_distance = dist
                min_index = i
        path.append(coordinates_list.pop(min_index))
    return path


def plot_points_and_path(coordinates_list, path, start_point):
    x_values, y_values = zip(*coordinates_list)
    path_x_values, path_y_values = zip(*path)

    plt.plot(x_values, y_values, 'o', color='lightgray', label='All Points')
    plt.plot(path_x_values, path_y_values, '-o', color='red', label='Path')
    plt.plot(start_point[0], start_point[1], 'o', color='purple', label='Start Point')
    plt.xlabel('X')
    plt.ylabel('Y')
    plt.title('Path')
    plt.grid(True)
    plt.legend()
    plt.show()


def write_waypoints_to_file(waypoints, filename):
    with open(filename, "w") as output_file:
        for counter, (x, y) in enumerate(waypoints, 1):
            output_file.write(f"waypoint:{counter}:{counter}:{x}:100:{y}:11:false:0:gui.xaero_default:false:0:false\n")


def create_index_excel(files):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Index"

    worksheet.cell(row=1, column=1, value="File Name")
    worksheet.cell(row=1, column=2, value="Number of Waypoints")

    for i, (file, waypoints) in enumerate(files.items(), 2):
        worksheet.cell(row=i, column=1, value=file)
        worksheet.cell(row=i, column=2, value=len(waypoints))

    return workbook


def get_file_path():
    if os.name == 'nt':  # Windows system
        from tkinter import Tk, filedialog
        Tk().withdraw()  # We don't want a full GUI, so keep the root window from appearing
        file_path = filedialog.askopenfilename(title="Select the input file",
                                               filetypes=(("Text files", "*.txt"), ("All files", "*.*")))
        return file_path
    else:  # Non-Windows system
        import readline
        def complete(text, state):
            results = [x for x in os.listdir('.') if x.startswith(text)] + [None]
            return results[state]

        readline.set_completer(complete)
        readline.parse_and_bind("tab: complete")
        return input("Please input the filename (use Tab for autocompletion): ")


def main():
    while True:
        start_x = input("Please input the X coordinate of the start point (default is 0): ")
        start_z = input("Please input the Z coordinate of the start point (default is 0): ")

        try:
            start_point = [int(start_x) if start_x else 0, int(start_z) if start_z else 0]
            break
        except ValueError:
            print("Invalid input. Please enter integer values for coordinates.")

    while True:
        try:
            num_paths = int(input("Please input the number of paths you want to explore: "))
            break
        except ValueError:
            print("Invalid input. Please enter a valid number.")

    while True:
        explore_ship_input = input(
            "Please input 'Y' if you want to explore the ship, or 'N' otherwise: ").strip().upper()
        if explore_ship_input in ('Y', 'N'):
            explore_ship = explore_ship_input == 'Y'
            break
        else:
            print("Invalid input. Please enter 'Y' or 'N'.")

    file_path = get_file_path()
    if not file_path:
        print("No file selected. Exiting.")
        return

    coordinates_list = []

    with open(file_path) as file:
        file.readline()  # Skip header
        for line in file:
            if "end_city" in line:
                _, _, x, y, details = line.split(";", 4)
                x, y = int(x), int(y)
                if explore_ship and "ship" in details or not explore_ship:
                    coordinates_list.append([x, y])
                    if explore_ship:
                        print(line)

    start_time = time.time()
    path = generate_path_module.generate_path(start_point, coordinates_list, num_paths)
    end_time = time.time()

    total_distance = sum(math.sqrt(calculate_distance_no_sqrt(path[i], path[i + 1])) for i in range(len(path) - 1))
    distance_ratio = total_distance / num_paths
    print(f"Total Distance: {total_distance:.2f}")
    print(f"Distance Ratio: {distance_ratio:.2f}")

    num_files = len(path) // 100 + (1 if len(path) % 100 != 0 else 0)

    files = {}
    for i in range(num_files):
        waypoints = path[i * 100: (i + 1) * 100]
        filename = f"output_{i + 1}.txt"
        files[filename] = waypoints
        write_waypoints_to_file(waypoints, filename)

    with zipfile.ZipFile("output.zip", "w") as zip_file:
        for file in files:
            zip_file.write(file)

    index_workbook = create_index_excel(files)
    index_workbook.save("index.xlsx")
    print(f"Execution time between breakpoints: {end_time - start_time:.6f} seconds")
    plot_points_and_path(coordinates_list, path, start_point)


if __name__ == "__main__":
    main()
